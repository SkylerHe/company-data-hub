#!/usr/bin/env python3
"""
option_report.py — export an option position as an Excel workbook you can chart.

Python does all the math (Greeks, IV, strategy P&L via volatility/option_strategy);
this writes it into a clean .xlsx so you can VIEW and DRAW charts in a spreadsheet
on a Mac — where IBKR's Windows-only Excel API isn't an option. Read-only: reads a
stored `option_quotes` snapshot (scrape it first) or computes offline from a spot+IV.

Sheets:
  Position — inputs, net premium, max profit/loss/break-even, net position Greeks.
  Greeks   — per-leg IV + delta/gamma/vega/theta/rho, and the position row.
  Payoff   — P&L across underlying prices at expiration + an embedded line chart
             (the payoff diagram); extend the range and the chart follows.

    # From a stored snapshot (run scrape_ibkr_options.py first):
    python option_report.py --strategy collar --company QQQ --expiry 2026-08-20 \
        --put-strike 713 --call-strike 754 --basis 700

    # Offline (no DB / no Gateway):
    python option_report.py --strategy collar --company QQQ --expiry 2026-08-20 \
        --put-strike 713 --call-strike 754 --spot 733.94 --iv 19.05 --days 7
"""

import argparse
import sys

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference

import option_strategy
import option_position as op
# reuse report.py's styling so the workbooks look like the valuation model
from report import _set, BOLD, TITLE, HEAD, HEAD_FILL, CALC_FILL, NOTE, NUM2, USD2

PAYOFF_SPAN = 0.30       # chart underlying from spot·(1±span)
PAYOFF_STEPS = 61        # rows in the payoff table


def _net_cash(strategy, legs):
    """Net premium: + = credit received, − = debit paid."""
    signs = dict(op.STRATEGY_LEGS[strategy])
    return sum(-signs.get(r, 0) * leg["premium"] for (r, _k), leg in legs.items())


def build_position(ws, ctx):
    a = option_strategy.analyze(ctx["strategy"], ctx["spot"], ctx["basis"],
                                ctx["kp"], ctx["cp"], ctx["kc"], ctx["cc"], ctx["contracts"])
    net = op.position_greeks(ctx["strategy"], ctx["legs"], ctx["contracts"])
    title = ctx["strategy"].replace("_", " ").title()
    _set(ws, "A1", f"{title} — {ctx['company']} {ctx['expiry']}", font=TITLE)
    _set(ws, "A2", f"Source: {ctx['source']}. Python-computed; edit/chart freely.", font=NOTE)

    rows = [
        ("INPUTS", None, None),
        ("Underlying spot", ctx["spot"], USD2),
        ("Share cost basis", ctx["basis"], USD2),
        ("Contracts (×100 sh)", ctx["contracts"], "0"),
        ("Net premium (+credit/−debit)", _net_cash(ctx["strategy"], ctx["legs"]), USD2),
        ("", None, None),
        ("P&L (per share · and total)", None, None),
    ]
    r = 4
    for label, val, fmt in rows:
        if val is None and fmt is None:
            _set(ws, f"A{r}", label, font=HEAD, fill=HEAD_FILL)
            _set(ws, f"B{r}", "", fill=HEAD_FILL)
        else:
            _set(ws, f"A{r}", label)
            _set(ws, f"B{r}", val, fmt=fmt)
        r += 1

    mult = a["mult"]
    def money(v):                       # unlimited stays as text
        return "unlimited" if v == float("inf") else ("−unlimited" if v == float("-inf") else v)
    for label, per in (("Max profit", a["max_profit"]), ("Max loss", a["max_loss"])):
        _set(ws, f"A{r}", label, font=BOLD)
        pv, tv = money(per), money(per * mult if abs(per) != float("inf") else per)
        _set(ws, f"B{r}", pv, fmt=USD2 if isinstance(pv, (int, float)) else None, font=BOLD)
        _set(ws, f"C{r}", tv, fmt=USD2 if isinstance(tv, (int, float)) else None)
        r += 1
    _set(ws, f"A{r}", "Break-even(s)", font=BOLD)
    _set(ws, f"B{r}", ", ".join(f"{b:g}" for b in a["break_evens"]) or "—"); r += 2

    _set(ws, f"A{r}", "NET POSITION GREEKS", font=HEAD, fill=HEAD_FILL)
    _set(ws, f"B{r}", f"× {mult} sh", font=HEAD, fill=HEAD_FILL); r += 1
    for g, unit in (("delta", "per $1"), ("gamma", "per $1²"), ("vega", "per +1 IV pt"),
                    ("theta", "per day"), ("rho", "per +1% rate")):
        _set(ws, f"A{r}", f"{g} ({unit})")
        _set(ws, f"B{r}", round(net[g], 3), fmt=NUM2, fill=CALC_FILL); r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    return a


def build_greeks(ws, ctx):
    _set(ws, "A1", "Per-leg Greeks (per share)", font=TITLE)
    hdr = ["Leg", "Side", "Strike", "IV", "Premium", "Delta", "Gamma", "Vega", "Theta", "Rho"]
    for j, h in enumerate(hdr):
        _set(ws, f"{chr(65+j)}3", h, font=HEAD, fill=HEAD_FILL)
    signs = dict(op.STRATEGY_LEGS[ctx["strategy"]])
    r = 4
    for right, k in ctx["needed"]:
        leg = ctx["legs"].get((right, k))
        if not leg:
            continue
        _set(ws, f"A{r}", f"{k:g}{right}")
        _set(ws, f"B{r}", "SELL" if signs.get(right, 0) < 0 else "BUY")
        _set(ws, f"C{r}", k, fmt=NUM2)
        _set(ws, f"D{r}", leg["iv"], fmt="0.0%")
        _set(ws, f"E{r}", leg["premium"], fmt=USD2)
        for j, g in enumerate(op.GREEKS):
            _set(ws, f"{chr(70+j)}{r}", round(leg[g], 5) if leg.get(g) is not None else None)
        r += 1

    net = op.position_greeks(ctx["strategy"], ctx["legs"], ctx["contracts"])
    _set(ws, f"A{r+1}", f"NET (× {100*ctx['contracts']} sh)", font=BOLD)
    for j, g in enumerate(op.GREEKS):
        _set(ws, f"{chr(70+j)}{r+1}", round(net[g], 3), font=BOLD, fill=CALC_FILL)
    for col in "ABCDEFGHIJ":
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["A"].width = 8


def build_payoff(ws, ctx):
    _set(ws, "A1", "Payoff at expiration", font=TITLE)
    _set(ws, "A2", "P&L if the underlying settles at each price. Chart is a line of "
                   "total P&L; extend the rows and it follows.", font=NOTE)
    for j, h in enumerate(("Underlying", "P&L / share", "P&L total")):
        _set(ws, f"{chr(65+j)}4", h, font=HEAD, fill=HEAD_FILL)
    lo, hi = ctx["spot"] * (1 - PAYOFF_SPAN), ctx["spot"] * (1 + PAYOFF_SPAN)
    step = (hi - lo) / (PAYOFF_STEPS - 1)
    mult = 100 * ctx["contracts"]
    first = 5
    for i in range(PAYOFF_STEPS):
        S = lo + i * step
        y = option_strategy._payoff(ctx["strategy"], S, ctx["spot"], ctx["basis"],
                                    ctx["kp"], ctx["cp"], ctx["kc"], ctx["cc"])
        row = first + i
        _set(ws, f"A{row}", round(S, 2), fmt=NUM2)
        _set(ws, f"B{row}", round(y, 2), fmt=USD2)
        _set(ws, f"C{row}", round(y * mult, 2), fmt=USD2)
    last = first + PAYOFF_STEPS - 1

    chart = LineChart()
    chart.title = "Payoff at expiration"
    chart.x_axis.title = "Underlying price"
    chart.y_axis.title = "P&L (total $)"
    chart.height, chart.width = 9, 18
    data = Reference(ws, min_col=3, min_row=4, max_row=last)         # incl header for title
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "E4")
    for col in "ABC":
        ws.column_dimensions[col].width = 13


def build_workbook(ctx):
    wb = Workbook()
    build_position(wb.active, ctx); wb.active.title = "Position"
    build_greeks(wb.create_sheet("Greeks"), ctx)
    build_payoff(wb.create_sheet("Payoff"), ctx)
    return wb


def main():
    ap = argparse.ArgumentParser(description="Export an option position as an Excel workbook")
    op.add_position_args(ap)
    ap.add_argument("--out", help="output .xlsx path")
    args = ap.parse_args()

    try:
        ctx = op.context_from_args(args)
    except ValueError as e:
        print(e, file=sys.stderr)
        sys.exit(1)
    out = args.out or f"{args.company}_{args.strategy}_options.xlsx"
    build_workbook(ctx).save(out)
    print(f"Wrote {out}  (sheets: Position · Greeks · Payoff — with a payoff chart)")


if __name__ == "__main__":
    main()
