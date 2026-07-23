#!/usr/bin/env python3
"""
report.py — export a valuation as an EDITABLE Excel model.

valuation.py prints a static read. This writes the same DCF as a live .xlsx where
the discounted-cash-flow math is real Excel formulas pointing at one editable
**Assumptions** sheet. Change the WACC, the growth rate, or beta in Excel and the
intrinsic value, upside, and margin of safety recompute — so the user can stress
the model themselves instead of trusting a single printed number. That's the
point for someone learning valuation: see how sensitive "intrinsic value" is to
assumptions you can't observe.

Sheets:
  Assumptions  — every input in one place; derived cost of equity + WACC as formulas.
  DCF          — a 3-scenario (bear/base/bull) projection grid; all formulas.
  Comps        — peer-median multiples pulled from this db (static snapshot).
  Summary      — verdict + intrinsic range; the headline number is a live link
                 to the DCF sheet, so it too moves with the assumptions.

    python report.py --company MSFT
    python report.py --company MSFT --out ~/Desktop/MSFT_valuation.xlsx
    python report.py --company MSFT --growth 0.10 --mos 0.30   # same overrides as valuation.py
"""

import argparse
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import analyze
import store
import valuation

# --- light styling ----------------------------------------------------------
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="305496")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow = editable
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")    # green  = formula
NOTE = Font(italic=True, size=9, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.0%"
USD = "$#,##0"
USD2 = "$#,##0.00"
NUM2 = "0.00"


def _set(ws, coord, value, *, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align)
    if border:
        c.border = BORDER
    return c


def build_assumptions(ws, d):
    """One editable input per row. Returns {key: 'Assumptions!$B$n'} so the DCF
    sheet can reference each by an absolute address in its formulas."""
    a, inp = d["assumptions"], d["inputs"]
    _set(ws, "A1", f"Valuation model — {d['company']}", font=TITLE)
    _set(ws, "A2", "Yellow cells are inputs you can edit — the DCF sheet recomputes.",
         font=NOTE)

    # (label, value, key, number_format).  key=None → a plain (non-referenced) row.
    scen = d["scenarios"]
    rows = [
        ("MARKET", None, None, None),
        ("Current price", d["price"], "price", USD2),
        ("Shares outstanding", inp["shares"], "shares", "#,##0"),
        ("Market cap", inp["market_cap"], "mktcap", USD),
        ("Net debt (debt − cash)", inp["net_debt"], "netdebt", USD),
        ("Total debt", inp["total_debt"], "debt", USD),
        ("", None, None, None),
        ("CASH FLOW", None, None, None),
        ("Latest FCF (starting point)", inp["fcf_latest"], "fcf0", USD),
        ("Historical FCF CAGR", inp["hist_fcf_cagr"], None, PCT),
        ("", None, None, None),
        ("DISCOUNT RATE", None, None, None),
        ("Risk-free rate", a["risk_free"], "rf", PCT),
        ("Beta", a["beta"], "beta", NUM2),
        ("Equity risk premium", a["equity_risk_premium"], "erp", PCT),
        ("Cost of debt (pre-tax)", a["cost_of_debt"], "kd", PCT),
        ("Tax rate", a["tax_rate"], "tax", PCT),
        ("", None, None, None),
        ("GROWTH", None, None, None),
        ("Stage-1 growth — BASE", scen["base"]["growth"], "base_g", PCT),
        ("Stage-1 growth — BEAR", scen["bear"]["growth"], "bear_g", PCT),
        ("Stage-1 growth — BULL", scen["bull"]["growth"], "bull_g", PCT),
        ("Terminal growth", a["terminal_growth"], "gterm", PCT),
        ("Forecast years", a["forecast_years"], "years", "0"),
        ("", None, None, None),
        ("Required margin of safety", a["mos_required"], "mos_req", PCT),
    ]

    ref = {}
    r = 4
    for label, value, key, fmt in rows:
        if label and value is None and key is None:  # section header
            _set(ws, f"A{r}", label, font=HEAD, fill=HEAD_FILL)
            _set(ws, f"B{r}", "", fill=HEAD_FILL)
        elif label:
            _set(ws, f"A{r}", label)
            editable = key is not None
            _set(ws, f"B{r}", value, fmt=fmt,
                 fill=INPUT_FILL if editable else None, border=editable)
            if key:
                ref[key] = f"Assumptions!$B${r}"
        r += 1

    # Derived rates (formulas, green) referencing the inputs above.
    r += 1
    _set(ws, f"A{r}", "DERIVED", font=HEAD, fill=HEAD_FILL)
    _set(ws, f"B{r}", "", fill=HEAD_FILL); r += 1
    _set(ws, f"A{r}", "Cost of equity  (rf + β·ERP)")
    _set(ws, f"B{r}", f"={ref['rf']}+{ref['beta']}*{ref['erp']}",
         fmt=PCT, fill=CALC_FILL, border=True)
    ref["ke"] = f"Assumptions!$B${r}"; r += 1
    _set(ws, f"A{r}", "WACC  (E/V·ke + D/V·kd·(1−tax))")
    v = f"({ref['mktcap']}+{ref['debt']})"
    _set(ws, f"B{r}",
         f"=IF({v}=0,{ref['ke']},{ref['mktcap']}/{v}*{ref['ke']}"
         f"+{ref['debt']}/{v}*{ref['kd']}*(1-{ref['tax']}))",
         fmt=PCT, fill=CALC_FILL, border=True)
    ref["wacc"] = f"Assumptions!$B${r}"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    return ref


def build_dcf(ws, d, ref):
    """3-scenario projection grid — every cash flow, discount factor, PV, terminal
    value, and per-share result is a live formula. Returns the base-case
    intrinsic cell so the Summary can link to it."""
    _set(ws, "A1", "Discounted cash flow — bear / base / bull", font=TITLE)
    _set(ws, "A2", "All values are formulas driven by the Assumptions sheet. "
                   "Edit assumptions, not these.", font=NOTE)

    n = int(d["assumptions"]["forecast_years"])
    scen_cols = {"bear": ("C", "D", ref["bear_g"]),
                 "base": ("E", "F", ref["base_g"]),
                 "bull": ("G", "H", ref["bull_g"])}

    # Header row
    hr = 4
    _set(ws, f"A{hr}", "Year", font=HEAD, fill=HEAD_FILL, align="center")
    _set(ws, f"B{hr}", "Disc. factor", font=HEAD, fill=HEAD_FILL, align="center")
    for name, (fc, pc, _) in scen_cols.items():
        _set(ws, f"{fc}{hr}", f"{name} FCF", font=HEAD, fill=HEAD_FILL, align="center")
        _set(ws, f"{pc}{hr}", f"{name} PV", font=HEAD, fill=HEAD_FILL, align="center")

    # Year 0 (starting FCF, no discounting)
    r0 = hr + 1
    _set(ws, f"A{r0}", 0, align="center")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{fc}{r0}", f"={ref['fcf0']}", fmt=USD)

    # Years 1..n
    first, last = r0 + 1, r0 + n
    for i in range(1, n + 1):
        r = r0 + i
        _set(ws, f"A{r}", i, align="center")
        _set(ws, f"B{r}", f"=1/(1+{ref['wacc']})^A{r}", fmt="0.000")
        for fc, pc, gref in scen_cols.values():
            _set(ws, f"{fc}{r}", f"={fc}{r-1}*(1+{gref})", fmt=USD)
            _set(ws, f"{pc}{r}", f"={fc}{r}*B{r}", fmt=USD)

    # Results block
    r = last + 2
    def label(txt):
        _set(ws, f"A{r}", txt, font=BOLD)

    label("PV of stage-1 FCF");
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"=SUM({pc}{first}:{pc}{last})", fmt=USD)
    pv_row = r; r += 1

    label("Terminal value (Gordon)")
    for fc, pc, _ in scen_cols.values():
        # FCF_N·(1+g) / (WACC−g)
        _set(ws, f"{pc}{r}",
             f"={fc}{last}*(1+{ref['gterm']})/({ref['wacc']}-{ref['gterm']})", fmt=USD)
    tv_row = r; r += 1

    label("PV of terminal value")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"={pc}{tv_row}*B{last}", fmt=USD)
    pvtv_row = r; r += 1

    label("Enterprise value")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"={pc}{pv_row}+{pc}{pvtv_row}", fmt=USD, font=BOLD)
    ev_row = r; r += 1

    label("Equity value  (EV − net debt)")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"={pc}{ev_row}-{ref['netdebt']}", fmt=USD)
    eq_row = r; r += 1

    label("Intrinsic value / share")
    intrinsic_cells = {}
    for name, (fc, pc, _) in scen_cols.items():
        _set(ws, f"{pc}{r}", f"={pc}{eq_row}/{ref['shares']}",
             fmt=USD2, font=BOLD, fill=CALC_FILL, border=True)
        intrinsic_cells[name] = f"DCF!{pc}{r}"
    ivps_row = r; r += 1

    label("Upside vs price  (intrinsic/price − 1)")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"={pc}{ivps_row}/{ref['price']}-1", fmt=PCT)
    r += 1

    label("Margin of safety  (1 − price/intrinsic)")
    for fc, pc, _ in scen_cols.values():
        _set(ws, f"{pc}{r}", f"=1-{ref['price']}/{pc}{ivps_row}", fmt=PCT)

    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
    return intrinsic_cells


def build_comps(ws, d):
    """Peer-median multiples snapshot (static — a cross-check on the DCF)."""
    _set(ws, "A1", "Comps — peer-median multiples", font=TITLE)
    c = d["comps"]
    _set(ws, "A2", f"Medians across {c['peers']} same-industry names in the db. "
                   "Applied to this company's trailing metric.", font=NOTE)
    if not c["values"]:
        _set(ws, "A4", "Not enough peers with clean multiples (need ≥3). "
                       "DCF stands alone.", font=NOTE)
        ws.column_dimensions["A"].width = 60
        return
    hdr = ["Multiple", "Peer median", "# peers", "Implied value / share"]
    for j, h in enumerate(hdr):
        _set(ws, f"{get_column_letter(j+1)}4", h, font=HEAD, fill=HEAD_FILL)
    r = 5
    for k, v in c["values"].items():
        _set(ws, f"A{r}", k)
        _set(ws, f"B{r}", v["multiple"], fmt=NUM2)
        _set(ws, f"C{r}", v["peers"])
        _set(ws, f"D{r}", v["implied_per_share"], fmt=USD2)
        r += 1
    _set(ws, f"A{r+1}", "EV/EBIT proxies EBITDA (EDGAR has no D&A) → conservative.",
         font=NOTE)
    for col, w in zip("ABCD", (22, 14, 10, 22)):
        ws.column_dimensions[col].width = w


def build_summary(ws, d, ref, intrinsic_cells):
    """Verdict + range. The headline intrinsic value LINKS to the live DCF cell,
    so the summary moves with the model."""
    _set(ws, "A1", f"Valuation summary — {d['company']}", font=TITLE)
    v = d["verdict"]
    rows = [
        ("Current price", f"={ref['price']}", USD2),
        ("Intrinsic value / share — BASE (live)", f"={intrinsic_cells['base']}", USD2),
        ("Intrinsic range — bear → bull (live)",
         f'={intrinsic_cells["bear"]}&"  →  "&{intrinsic_cells["bull"]}', None),
        ("Margin of safety (live)",
         f"=1-{ref['price']}/{intrinsic_cells['base']}", PCT),
        ("Required margin of safety", f"={ref['mos_req']}", PCT),
    ]
    r = 3
    for label, formula, fmt in rows:
        _set(ws, f"A{r}", label, font=BOLD)
        _set(ws, f"B{r}", formula, fmt=fmt, fill=CALC_FILL, border=True)
        r += 1

    r += 1
    _set(ws, f"A{r}", "VERDICT", font=HEAD, fill=HEAD_FILL)
    _set(ws, f"B{r}", v["call"], font=HEAD, fill=HEAD_FILL); r += 1
    _set(ws, f"A{r}", "Rationale", font=BOLD)
    _set(ws, f"B{r}", v["reason"]); r += 2

    rv = d["reverse_dcf"]["implied_stage1_growth"]
    _set(ws, f"A{r}", "Reverse DCF — growth the price implies", font=BOLD)
    _set(ws, f"B{r}", rv if rv is not None else "outside plausible range",
         fmt=PCT if rv is not None else None); r += 1
    _set(ws, f"A{r}", "…vs historical FCF CAGR", font=BOLD)
    _set(ws, f"B{r}", d["reverse_dcf"]["hist_fcf_cagr"], fmt=PCT); r += 2

    _set(ws, f"A{r}", "This is a model, not a fact. The verdict is the base case at "
                      "the Assumptions-sheet inputs; edit them to test it.", font=NOTE)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30


def build_workbook(d):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_assume = wb.create_sheet("Assumptions")
    ws_dcf = wb.create_sheet("DCF")
    ws_comps = wb.create_sheet("Comps")

    ref = build_assumptions(ws_assume, d)
    intrinsic_cells = build_dcf(ws_dcf, d, ref)
    build_comps(ws_comps, d)
    build_summary(ws_sum, d, ref, intrinsic_cells)
    return wb


def main():
    ap = argparse.ArgumentParser(description="Export a valuation as an editable Excel model")
    ap.add_argument("--db", default=store.DB_PATH)
    ap.add_argument("--company", required=True, help="company name or ticker")
    ap.add_argument("--out", help="output .xlsx path (default <TICKER>_valuation.xlsx)")
    # Same assumption overrides as valuation.py so the model can be seeded already-stressed.
    ap.add_argument("--growth", type=float)
    ap.add_argument("--rf", type=float)
    ap.add_argument("--erp", type=float)
    ap.add_argument("--cost-of-debt", type=float, dest="cost_of_debt")
    ap.add_argument("--tax", type=float, dest="tax_rate")
    ap.add_argument("--terminal", type=float, dest="terminal_growth")
    ap.add_argument("--years", type=int, dest="forecast_years")
    ap.add_argument("--mos", type=float, dest="mos_required")
    args = ap.parse_args()

    overrides = {k: v for k, v in vars(args).items()
                 if k in valuation.DEFAULTS and v is not None}
    if args.growth is not None:
        overrides["growth"] = args.growth

    db = store.get_db(args.db)
    name = analyze.resolve_company(db, args.company)
    if not name:
        print(f"'{args.company}' not found in {args.db}", file=sys.stderr)
        sys.exit(1)

    d = valuation.value_company(db, name, overrides)
    if not d["dcf"]["intrinsic_per_share"]:
        print(f"Can't build a DCF model for {name} — no positive FCF/shares. "
              f"Refresh via company-data first.", file=sys.stderr)
        sys.exit(1)

    ticker_row = db.execute("SELECT ticker FROM instruments WHERE name=?", (name,)).fetchone()
    ticker = (ticker_row[0] if ticker_row and ticker_row[0] else name).replace(" ", "_")
    out = args.out or f"{ticker}_valuation.xlsx"

    build_workbook(d).save(out)
    print(f"Wrote {out}")
    print(f"  Summary verdict: {d['verdict']['call']} — base intrinsic "
          f"${d['dcf']['intrinsic_per_share']:,.2f} vs price ${d['price']:,.2f}")
    print("  Open it and edit the yellow Assumptions cells to stress the model.")
    db.close()


if __name__ == "__main__":
    main()
